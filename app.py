import gradio as gr
from PIL import Image
import torch
import torchvision.models as models
import torchvision.transforms as transforms
import cv2
import numpy as np
import openpyxl
import os
from tkinter import filedialog

# Load the pre-trained EfficientNet-B7 model
model = models.efficientnet_b7(pretrained=True)
model.eval()

# Define the transformations to be applied to the input image
transform = transforms.Compose([
    transforms.Resize((224, 224)),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406],
                         std=[0.229, 0.224, 0.225])
])
def predict_house_area(room_id, excel_file, image_files):
    total_area_sqm = 0
    predicted_areas = []

    # Check if the excel_file is provided
    if excel_file is not None:
        # Load the existing Excel workbook
        workbook = openpyxl.load_workbook(excel_file.name)
        worksheet = workbook.active
    else:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Write the headers to the worksheet
        worksheet.cell(row=1, column=1).value = "Room ID"
        worksheet.cell(row=1, column=2).value = "Image File"
        worksheet.cell(row=1, column=3).value = "Predicted Area (sqm)"

    # Get the last row index to append new data
    last_row_index = worksheet.max_row if worksheet.max_row else 1

    # Loop over all the images
    for i, image_file in enumerate(image_files):
        # Load the input image
        img = Image.open(image_file.name)
        # Extract the image file name from the path
        image_file_name = os.path.basename(image_file.name)
        # Check if the image is PNG and convert to JPEG if it is
        if img.format == "PNG":
            # Convert the image to RGB format
            img = img.convert("RGB")

        # Apply the transformations to the input image
        img_transformed = transform(img)

        # Add a batch dimension to the transformed image tensor
        img_transformed_batch = torch.unsqueeze(img_transformed, 0)

        # Use the pre-trained model to make a prediction on the input image
        with torch.no_grad():
            output = model(img_transformed_batch)

        # Convert the output tensor to a probability distribution using softmax
        softmax = torch.nn.Softmax(dim=1)
        output_probs = softmax(output)

        # Extract the predicted class (house square footage) from the output probabilities
        predicted_class = torch.argmax(output_probs)

        # Calculate the predicted area based on the predicted class
        predicted_area_sqm = 0
        if predicted_class in [861, 648, 594, 894, 799, 896, 454]:
            # Convert to grayscale and apply adaptive thresholding
            gray = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2GRAY)
            gray = cv2.GaussianBlur(gray, (5, 5), 0)
            mask = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 11, 2)

            # Apply Canny edge detection to the binary mask
            edges = cv2.Canny(mask, 30, 100)

            # Apply dilation to fill gaps in the contour
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
            dilated = cv2.dilate(edges, kernel, iterations=2)
            eroded = cv2.erode(dilated, kernel, iterations=1)

            # Find contours in binary mask
            contours, _ = cv2.findContours(eroded, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

            # Find largest contour and calculate area
            max_area = 0
            for c in contours:
                area = cv2.contourArea(c)
                if area > max_area:
                    max_area = area

            # Convert pixel area to square meters
            pixels_per_meter = 300  # adjust this value based on your image resolution and actual room dimensions
            predicted_area_sqm = (max_area + 10) / (2 * pixels_per_meter ** 2)
        else:
            predicted_area_sqft = predicted_class.item()
            predicted_area_sqm = predicted_area_sqft * 0.092903 / 4.2

        # Add the predicted area to the sum
        total_area_sqm += predicted_area_sqm

        # Add the predicted area to the list of predicted areas
        predicted_areas.append(predicted_area_sqm)

        # Write the room ID, image file name, and predicted area to the worksheet
        worksheet.cell(row=last_row_index + i + 1, column=1).value = room_id
        worksheet.cell(row=last_row_index + i + 1, column=2).value = image_file_name
        worksheet.cell(row=last_row_index + i + 1, column=3).value = predicted_area_sqm

    # Save the workbook to a temporary file
    temp_file = "predicted_areas.xlsx"
    workbook.save(temp_file)

    # Get the path of the first uploaded image
    first_image_path = image_files[0].name if image_files else None

    return f"Sum of predicted house square footage: {total_area_sqm:.2f} square meters", temp_file ,first_image_path


inputs = [
    gr.inputs.Textbox(label = "Mã Phòng" , type = "text"),
    gr.inputs.File(label="Excel File", type="file"),
    gr.inputs.File(label="Images", type="file", file_count="multiple")
]

outputs = [
    gr.outputs.Textbox(label="Sum of Predicted House Square Footage"),
    gr.outputs.File(label="Excel Result"),
    gr.outputs.Image(type="pil", label="Uploaded Image")
]

interface = gr.Interface(
    fn=predict_house_area,
    inputs=inputs,
    outputs=outputs,
    title="House Predictor",
    allow_flagging="never"  # Disable flag button
)

if __name__ == "__main__":
    interface.launch()

